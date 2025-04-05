import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Define column names based on SQL schema
columns = [
    "id", "name", "price", "mrp", "discount", "category", "manufacturer",
    "pack_size", "rating", "reviews", "image", "prescription_required", "in_stock",
    "description", "features", "benefits", "usage", "ingredients",
    "safety_information", "safety_info", "storage", "quantity", "expiry_date",
    "weight", "pack_form", "country_of_origin", "related_products"
]

# Create an empty DataFrame
df = pd.DataFrame(columns=columns)

# Save DataFrame to Excel
excel_file = "Products.xlsx"
df.to_excel(excel_file, index=False, engine="openpyxl")

# Load workbook for formatting
wb = Workbook()
ws = wb.active
ws.title = "Products"

# Write headers with styling
for col_num, column_title in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_num, value=column_title)
    cell.font = Font(bold=True)  # Make header bold
    ws.column_dimensions[get_column_letter(col_num)].width = max(len(column_title) + 2, 15)  # Auto-adjust width

# Save the formatted workbook
wb.save(excel_file)

print(f" Excel file '{excel_file}' created successfully with proper formatting!")
